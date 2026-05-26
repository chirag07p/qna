import io
import os
import pandas as pd
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import our matcher engine
from matcher import match_sheets

app = FastAPI(title="Excel Q&A Matcher API", version="1.0.0")

# Enable CORS so frontend can talk to backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Paths for storing uploaded files locally
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)

TEMP_FILE1_PATH = os.path.join(DATA_DIR, "temp_questions.xlsx")
TEMP_FILE2_PATH = os.path.join(DATA_DIR, "temp_answers.xlsx")

@app.post("/api/upload")
async def upload_files(
    file1: Optional[UploadFile] = File(None),
    file2: Optional[UploadFile] = File(None)
):
    """
    Uploads questions (file1) and reference answers (file2) sheets.
    Returns column names and first 5 preview rows for configuration.
    If no files are uploaded, falls back to pre-placed files in server/data/.
    """
    df1 = None
    if file1 is not None and file1.filename:
        contents = await file1.read()
        try:
            if file1.filename.endswith(".csv"):
                df1 = pd.read_csv(io.BytesIO(contents))
            else:
                df1 = pd.read_excel(io.BytesIO(contents))
            # Save as Excel locally for subsequent matching
            df1.to_excel(TEMP_FILE1_PATH, index=False)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parsing File 1: {str(e)}")
    else:
        # Check if pre-placed file exists
        local_path = os.path.join(DATA_DIR, "questions.xlsx")
        if os.path.exists(local_path):
            try:
                df1 = pd.read_excel(local_path)
                # Copy to temp path
                df1.to_excel(TEMP_FILE1_PATH, index=False)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error reading local questions.xlsx: {str(e)}")
        elif os.path.exists(TEMP_FILE1_PATH):
            try:
                df1 = pd.read_excel(TEMP_FILE1_PATH)
            except Exception as e:
                pass

    df2 = None
    if file2 is not None and file2.filename:
        contents = await file2.read()
        try:
            if file2.filename.endswith(".csv"):
                df2 = pd.read_csv(io.BytesIO(contents))
            else:
                df2 = pd.read_excel(io.BytesIO(contents))
            # Save as Excel locally
            df2.to_excel(TEMP_FILE2_PATH, index=False)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parsing File 2: {str(e)}")
    else:
        # Check if pre-placed file exists
        local_path = os.path.join(DATA_DIR, "answers.xlsx")
        if os.path.exists(local_path):
            try:
                df2 = pd.read_excel(local_path)
                # Copy to temp path
                df2.to_excel(TEMP_FILE2_PATH, index=False)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error reading local answers.xlsx: {str(e)}")
        elif os.path.exists(TEMP_FILE2_PATH):
            try:
                df2 = pd.read_excel(TEMP_FILE2_PATH)
            except Exception as e:
                pass

    if df1 is None:
        raise HTTPException(status_code=400, detail="Question sheet (File 1) is missing. Please upload a file.")
    if df2 is None:
        raise HTTPException(status_code=400, detail="Reference Q&A sheet (File 2) is missing. Please upload a file.")

    # Generate previews (limiting rows and converting NaN to None for JSON serialization)
    def clean_preview(df):
        preview_rows = df.head(5).copy()
        # Replace NaN with None
        preview_rows = preview_rows.where(pd.notnull(preview_rows), None)
        return preview_rows.to_dict(orient="records")

    return {
        "file1_columns": list(df1.columns),
        "file2_columns": list(df2.columns),
        "file1_preview": clean_preview(df1),
        "file2_preview": clean_preview(df2),
    }

class MatchRequest(BaseModel):
    cname1: str
    cname2: str
    ans_cname: str
    threshold: float
    top_k: int = 3

@app.post("/api/match")
async def match_endpoint(req: MatchRequest):
    """
    Runs the hybrid matching engine using choices from the user.
    """
    if not os.path.exists(TEMP_FILE1_PATH) or not os.path.exists(TEMP_FILE2_PATH):
        raise HTTPException(status_code=400, detail="Files not found. Please upload them first.")

    try:
        sheet1 = pd.read_excel(TEMP_FILE1_PATH)
        sheet2 = pd.read_excel(TEMP_FILE2_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading temporary data: {str(e)}")

    try:
        results = match_sheets(
            sheet1=sheet1,
            sheet2=sheet2,
            cname1=req.cname1,
            cname2=req.cname2,
            ans_cname=req.ans_cname,
            t=req.threshold,
            top_k=req.top_k
        )
        return {"results": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Matching calculation failed: {str(e)}")

class MatchItem(BaseModel):
    matched_question: str
    answer: str
    score: float

class MatchResultRow(BaseModel):
    row_index: int
    original_question: str
    matches: List[MatchItem]
    is_conflict: bool
    selected_match_idx: Optional[int]
    original_row: Dict[str, Any]

class ExportRequest(BaseModel):
    export_mode: str  # "best", "multi", "concatenated"
    cname1: str       # Question column name in Sheet 1
    ans_cname: str    # Target Answer column name
    results: List[MatchResultRow]

@app.post("/api/export")
async def export_endpoint(req: ExportRequest):
    """
    Generates and streams a styled Excel sheet with matched answers.
    """
    if not req.results:
        raise HTTPException(status_code=400, detail="No matching results provided for export.")

    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matched Results"

    # Set gridlines to visible
    ws.views.sheetView[0].showGridLines = True

    # 1. Define styling tokens
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Elegant Slate Dark Gray
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    body_font = Font(name=font_family, size=10)
    conflict_font = Font(name=font_family, size=10, color="991B1B", bold=True)
    
    # Soft confidence scoring color fills
    high_conf_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # soft light green
    med_conf_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")   # soft light yellow
    low_conf_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # soft light red
    
    border_side = Side(border_style="thin", color="E5E7EB")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # 2. Determine and write the header row
    # Reconstruct headers from original_row keys in the first result
    first_row_orig = req.results[0].original_row
    original_cols = list(first_row_orig.keys())
    
    headers = []
    # Include all original columns
    headers.extend(original_cols)
    
    # Add new output columns depending on export mode
    if req.export_mode == "best":
        headers.extend(["Matched Answer", "Confidence Score (%)", "Status"])
    elif req.export_mode == "multi":
        # Find maximum number of matches returned across rows
        max_matches = 0
        for r in req.results:
            max_matches = max(max_matches, len(r.matches))
        # Ensure we add columns up to max_matches
        for idx in range(1, max_matches + 1):
            headers.extend([f"Answer Candidate {idx}", f"Confidence {idx} (%)"])
        headers.append("Status")
    elif req.export_mode == "concatenated":
        headers.extend(["All Matched Answers", "Status"])
    else:
        raise HTTPException(status_code=400, detail=f"Invalid export mode: {req.export_mode}")

    # Write header row
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 28

    # 3. Write data rows
    for r_idx, r in enumerate(req.results, start=2):
        row_data = []
        
        # Add original row values
        for col in original_cols:
            row_data.append(r.original_row.get(col, ""))
            
        # Get selected match or matches
        selected_idx = r.selected_match_idx
        matches = r.matches
        
        status_label = "Conflict" if r.is_conflict else ("Matched" if matches else "No Match")
        
        # Add values based on export mode
        if req.export_mode == "best":
            best_ans = ""
            best_score = ""
            if selected_idx is not None and 0 <= selected_idx < len(matches):
                best_ans = matches[selected_idx].answer
                best_score = matches[selected_idx].score
            elif matches:
                best_ans = matches[0].answer
                best_score = matches[0].score
                
            row_data.extend([best_ans, best_score, status_label])
            
        elif req.export_mode == "multi":
            max_matches = 0
            for row in req.results:
                max_matches = max(max_matches, len(row.matches))
                
            for idx in range(max_matches):
                if idx < len(matches):
                    row_data.extend([matches[idx].answer, matches[idx].score])
                else:
                    row_data.extend(["", ""])
            row_data.append(status_label)
            
        elif req.export_mode == "concatenated":
            concat_str = ""
            if matches:
                concat_items = []
                for idx, m in enumerate(matches, start=1):
                    concat_items.append(f"{idx}. [{m.score}%] {m.answer}")
                concat_str = "\n".join(concat_items)
            row_data.extend([concat_str, status_label])

        ws.append(row_data)
        ws.row_dimensions[r_idx].height = 24
        
        # Style the cells in the row
        curr_row = list(ws.iter_rows(min_row=r_idx, max_row=r_idx))[0]
        
        for c_idx, cell in enumerate(curr_row, start=1):
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = left_align
            
            # Highlight status column
            col_name = headers[c_idx - 1]
            if col_name == "Status":
                cell.alignment = center_align
                if cell.value == "Conflict":
                    cell.font = conflict_font
                    cell.fill = low_conf_fill
                elif cell.value == "No Match":
                    cell.fill = low_conf_fill
                else:
                    cell.fill = high_conf_fill
                    
            # Color confidence scores
            if "Confidence" in col_name or col_name == "Confidence Score (%)":
                cell.alignment = center_align
                if cell.value is not None and cell.value != "":
                    try:
                        val = float(cell.value)
                        if val >= 80:
                            cell.fill = high_conf_fill
                        elif val >= 50:
                            cell.fill = med_conf_fill
                        else:
                            cell.fill = low_conf_fill
                    except:
                        pass

    # 4. Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            # If multiple lines (e.g. concatenated), take longest line
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        # Cap width to avoid extremely wide columns
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    # 5. Stream the Excel file back
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"matched_results_{req.export_mode}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    